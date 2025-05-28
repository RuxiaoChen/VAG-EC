"""
Train an RNN decoder to make binary predictions;
then train an RNN language model to generate sequences
"""


import contextlib
import csv
from collections import defaultdict
import pdb
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from models import my_GCN, combiner, custom_loss
from torch_geometric.nn import SAGEConv, to_hetero
import models
import util
import data
import os
import vis
import emergence
import torch.nn.functional as F
from openpyxl import Workbook, load_workbook
import pandas as pd
import io_util

# Logging
import logging


# channel impact
from channel_impact import *
# import prepare_listener_data as ld
logging.basicConfig(
    format="%(asctime)s %(levelname)-8s %(message)s",
    level=logging.INFO,
    datefmt="%Y-%m-%d %H:%M:%S",
)


def convert_lang_to_numeric(lang, lang_length, pad_val=-1, skip_sos_eos=True):
    """
    Convert lang to numeric, with custom padding, for later language analysis
    """
    lang_i = lang.argmax(2)
    for i, length in enumerate(lang_length):

        if skip_sos_eos:
            # zero out EOS
            lang_i[i, length - 1 :] = pad_val
        else:
            lang_i[i, length:] = pad_val

    # shave off SOS, ending EOS if present
    if skip_sos_eos:
        lang_i = lang_i[:, 1:-1]

    return lang_i


def get_true_lang(batch, dataset, args, join=True):
    spk_inp, spk_y, lis_inp, lis_y, names, parts_img, all_edges, true_lang, md, idx = batch
    true_lang_text = dataset.to_text(true_lang, join=join)
    return true_lang_text


def get_positive_examples(inp, y):
    """
    inp -> batch_size x n_examples x feat_size
    y -> batch_size x y

    output
    """
    where_zero = np.where(y.sum(1) == 0)[0]
    y[where_zero] = 1
    occur_rows, occur_cols = np.where(y)
    row_indices, occur_col_indices = np.unique(occur_rows, return_index=True)
    assert (row_indices == np.arange(len(row_indices))).all()
    assert len(occur_col_indices) == len(y)
    col_indices = occur_cols[occur_col_indices]
    sel = inp[row_indices, col_indices]
    return sel


def subsample(items, idx):
    return [items[i] for i in idx]


def compute_lang_metrics(
    all_lang,
    dataset,
    args,
    attrs=None,
    reprs=None,
    attrs_numeric=None,
    toks=None,
    max_analysis_length=1000,
):
    lang_metrics = {}
    if all_lang.shape[0] > max_analysis_length:
        idx = np.random.choice(
            all_lang.shape[0], size=max_analysis_length, replace=False
        )

        all_lang = all_lang.iloc[idx].reset_index()

        if attrs is not None:
            attrs = subsample(attrs, idx)
        if toks is not None:
            toks = subsample(toks, idx)
        if reprs is not None:
            reprs = subsample(reprs, idx)

    # topographic similarity between ground truth language and tokens
    # only do it if the ground truth language is meaningful
    if dataset.name == "shapeworld":
        langts = emergence.topsim(
            all_lang["true_lang"], toks, meaning_distance_fn="edit"
        )
        lang_metrics["langts"] = langts

    if dataset.name == "shapeworld":

        def compute_hd(tl1, tl2):
            # Remove SOS, EOS
            tl1 = " ".join(tl1[1:-1])
            tl2 = " ".join(tl2[1:-1])
            return dataset.concept_distance(tl1, tl2)

    elif dataset.name == "cub":

        def compute_hd(tl1, tl2):
            tl1 = int(tl1[1])
            tl2 = int(tl2[1])
            return dataset.concept_distance(tl1, tl2)

    if dataset.concept_distances is not None:
        # hd = emergence.topsim(
        #     all_lang["true_lang"], toks, meaning_distance_fn=dataset.concept_distances      # compute_hd
        # )
        hd=0
        lang_metrics["hausdorff"] = hd

    if attrs is not None:
        # topographic similarity between meanings and tokens
        ts = emergence.topsim(
            attrs, toks, meaning_distance_fn=dataset.meaning_distance_fn
        )
        lang_metrics["ts"] = ts

        # topographic similarity between reprs and attributes
        # For random sets later, worth disentangling prototype repr from
        # individual inputs repr
        reprts = emergence.topsim(
            attrs,
            reprs,
            meaning_distance_fn=dataset.meaning_distance_fn,
            message_distance_fn="euclidean",
        )
        lang_metrics["reprts"] = reprts

    return lang_metrics


def compute_metrics_by_md(all_lang, md_vocab=None):
    metrics_by_md = {}
    per_md_acc = all_lang[["md", "acc"]].groupby("md").mean()
    for i, md_row in per_md_acc.iterrows():
        if md_vocab is None:
            md_name = str(md_row.name)
        else:
            md_name = md_vocab["i2w"][md_row.name]
        md_key = f"acc_md_{md_name}"
        metrics_by_md[md_key] = md_row["acc"]
    return metrics_by_md


def log_epoch_summary(epoch, split, metrics):
    logging.info(
        "Epoch {}\t{} {}".format(
            epoch,
            split.upper(),
            " ".join("{}: {:.4f}".format(m, v) for m, v in metrics.items()),
        )
    )


def log_epoch_progress(epoch, batch_i, batch_size, dataloader, stats):
    meter_str = " ".join(f"{k}: {v.avg:.3f}" for k, v in stats.meters.items())
    data_i = batch_i * batch_size
    data_total = len(dataloader.dataset)
    pct = round(100 * batch_i / len(dataloader))
    logging.info(f"Epoch {epoch} [{data_i}/{data_total} ({pct}%)] {meter_str}")


def init_metrics():
    """
    Initialize the metrics for this training run. This is a defaultdict, so
    metrics not specified here can just be appended to/assigned to during
    training.
    Returns
    -------
    metrics : `collections.defaultdict`
        All training metrics
    """
    metrics = {}
    metrics["best_acc"] = 0.0
    metrics["best_val_acc"] = 0.0
    metrics["best_val_same_acc"] = 0.0
    metrics["best_loss"] = float("inf")
    metrics["best_epoch"] = 0
    return metrics

def gcn_outfea(out):
    food_lis=out['food']
    food_lis=food_lis.view(1, -1)[0]
    tool_lis=out['tableware']
    tool_lis=tool_lis.view(1, -1)[0]
    final_feavec = torch.cat((food_lis, tool_lis), dim=0)
    return final_feavec

def append_list_to_csv(file_path, string_list):
    # 打开文件，以追加模式写入
    with open(file_path, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        # 将列表作为一行写入文件
        writer.writerow(string_list)

def run(
    epsil,
    split,
    epoch,
    pair,
    optimizer,
    dataloaders,
    args,
    random_state=None,
    force_no_train=False,
):
    """
    Run the model for a single epoch.

    Parameters
    ----------
    split : ``str``
        The dataloader split to use. Also determines model behavior if e.g.
        ``split == 'train'`` then model will be in train mode/optimizer will be
        run.
    epoch : ``int``
        current epoch
    model : ``torch.nn.Module``
        the model you are training/evaling
    optimizer : ``torch.nn.optim.Optimizer``
        the optimizer
    criterion : ``torch.nn.loss``
        the loss function
    dataloaders : ``dict[str, torch.utils.data.DataLoader]``
        Dictionary of dataloaders whose keys are the names of the ``split``s
        and whose values are the corresponding dataloaders
    args : ``argparse.Namespace``
        Arguments for this experiment run
    random_state : ``np.random.RandomState``
        The numpy random state in case anything stochastic happens during the
        run

    Returns
    -------
    metrics : ``dict[str, float]``
        Metrics from this run; keys are statistics and values are their average
        values across the batches
    """

    # epsil = 0    # 误码率
    print(split)
    training = (split == "train") and not force_no_train
    dataloader = dataloaders[split]
    torch.set_grad_enabled(training)  # 用于开启或关闭梯度计算

    pair.train(mode=True)

    stats = util.Statistics()

    all_lang = []
    all_toks = []  # language, unjoined text form, ragged
    # FIXME - make this one class
    if dataloader.dataset.name == "cub":
        all_attrs = []
        all_reprs = []  # representations
    else:
        all_attrs = None
        all_reprs = None

    # 这段代码主要用于训练过程中的一些准备工作，如清零梯度和根据训练轮数计算一些参数值。这些参数可能是在训练过程中需要进行调整的超参数。
    if training:  # sample.py 未执行这个if，执行了下面的else  # 实际训练时执行
        optimizer.zero_grad()    # 在每个训练步骤之前清空模型参数的梯度，以避免上一次训练的梯度对本次训练的影响
        for opt in gcn_optimizer:
            opt.zero_grad()
        this_epoch_eps = max(0.0, args.eps - (epoch * args.eps_anneal))
        this_epoch_uniform_weight = max(
            0.0, args.uniform_weight - (epoch * args.uniform_weight_anneal)
        )
        this_epoch_softmax_temp = max(
            1.0, args.softmax_temp - (epoch * args.softmax_temp_anneal)
        )
    else:
        this_epoch_eps = 0.0
        this_epoch_uniform_weight = 0.0    # 测试集
        this_epoch_softmax_temp = 1.0
    all_messages = []
    all_states = []
    for batch_i, batch in enumerate(dataloader):
        torch.cuda.empty_cache()
        # print('now is batch:', batch_i)
        spk_inp, spk_y, lis_inp, lis_y, names, inp_parts, all_edges, true_lang, md, idx = batch  # lis_inp size: 4*6*3*224*224
        file_path=f'path_{args.vocab_size}_{args.max_lang_length}.csv'
        string_list=names[0][0]
        append_list_to_csv(file_path, string_list)
        batch_size = spk_inp.shape[0]

        # Determine what's input
        if dataloader.dataset.name == "shapeworld":
            spk_inp = spk_inp.float() / 255
            # lis_inp = lis_inp.float() / 255
        else:
            spk_inp = spk_inp.float()
            inp_parts = inp_parts.float()
            # lis_inp = lis_inp.float()
        spk_y = spk_y.float()
        lis_y = lis_y.float()
        # 将数据移动到CUDA设备上，利用GPU的并行计算能力加速模型的训练
        if args.cuda:
            spk_inp = spk_inp.cuda()
            spk_y = spk_y.cuda()
            # lis_inp = lis_inp.cuda()
            lis_y = lis_y.cuda()
            inp_parts = inp_parts.cuda()
            spk_inp_parts = inp_parts[:, 0:1, :, :, :]
            spk_all_edges = [all_edges[0]]

            (lang1, lang_length), states = pair.speaker(
                spk_inp,
                spk_y,
                spk_inp_parts,
                spk_all_edges,
                model_gcn_s,
                attention_s,
                max_len=args.max_lang_length,
                eps=this_epoch_eps,
                softmax_temp=this_epoch_softmax_temp,
                uniform_weight=this_epoch_uniform_weight,
            )

            # 这里可以加现实信道
            all_messages.append(lang1.cpu())    # 把经过噪声前的消息储存下来
            all_states.append(states.cpu())
            lang1=message_in_channel2(lang1,epsil)

            # 原始signaling game
            # lis_scores = pair.listener(spk_inp, lang1, lang_length)
            # print(lis_scores)

            # for i in range(0, batch_size):
            #     new_names.append([names[0][0][i]])
            #     new_names.append([n[i] for n in names[1]])
            #     all_names.append(new_names)
            #     new_names=[]
            #
            # for names1 in all_names:
            #     for n in names1:
            #         for idxx in n:
            #             if isinstance(idxx, np.ndarray):
            #                 idxx = 'data/' + idxx[0]
            #             else:
            #                 idxx = 'data/' + idxx
            #             idxx = idxx[:-4]
            #             # lspath = idxx[:12] + '_graph_2' + idxx[20:] + '/'
            #             lspath = idxx[:12] + '_graph' + idxx[12:] + '/'
            #             # lspath = lspath[:5]+'my_food_graph/'+lspath[5:]
            #             lis_one_inp = ld.get_lis_data(lspath)
            #             lis_ginp.append(lis_one_inp)    # store heter data in list
            #     all_lis_ginp.append(lis_ginp)
            #     lis_ginp=[]

            # 把信息传送给listener
            # all_lis_gra_inp=[]
            # new_inp=[]
            # for a_lis in all_lis_ginp:
            #     for b_lis in a_lis:
            #         edge = torch.cat(list(b_lis.edge_index_dict.values()), dim=1).cuda()
            #         aaaa = torch.cat(( b_lis.x_dict['tableware'], b_lis.x_dict['food'], b_lis.x_dict['other'], b_lis.x_dict['relation']), dim=0)
            #         # pdb.set_trace()
            #         # out = model_gcn(aaaa, edge)
            #         # if out.size(0) > 33:
            #         #     out = out[:33]
            #         # elif out.size(0) < 33:
            #         #     mean_tensor = torch.mean(out, dim=0, keepdim=True)
            #         #     out = torch.cat([out, mean_tensor.repeat(33 - out.size(0), 1)],
            #         #                              dim=0)
            #         aaaa = model_gin_t(aaaa, edge)
            #         global_mean_pool = torch.mean(aaaa, dim=0)
            #         # attention_weights = attention(aaaa)
            #         # # 应用注意力权重
            #         # weighted_features = attention_weights * aaaa
            #         # # 全局特征汇聚
            #         # global_feature = torch.sum(weighted_features, dim=0).cuda()
            #         new_inp.append(global_mean_pool)
            #     all_lis_gra_inp.append(new_inp)
            #     new_inp=[]
            # all_lis_gra_inp = torch.stack([torch.stack(inner_list) for inner_list in all_lis_gra_inp])
            # all_lis_gra_inp.to(device)

            lis_scores, img_embedding= pair.listener(inp_parts, all_edges, model_gcn_l, attention_l, lang1, lang_length)

            # lis_scores= pair.listener(spk_inp, all_edges, model_gin_v, lang1, lang_length)

        # Evaluate loss and accuracy
        if args.reference_game_xent:
            # Take only 0th listener score + after midpoint. Then do cross entropy
            # assert lis_scores.shape[1] % 2 == 0
            # midp = lis_scores.shape[1] // 2
            # lis_scores_xent = torch.cat((lis_scores[:, :1], lis_scores[:, midp:]), 1)
            zeros = torch.zeros(batch_size, dtype=torch.int64, device=lis_scores.device)
            this_loss = pair.xent_criterion(lis_scores, zeros)
            lis_pred = lis_scores.argmax(1)
            per_game_acc = (lis_pred == 0).float().cpu().numpy()
            this_acc = per_game_acc.mean()
            # print('loss',this_loss)
            # print('accuracy:',this_acc)
        else:   # sample.py执行
            this_loss = pair.bce_criterion(lis_scores, lis_y)    # 计算二元交叉熵损失
            # this_loss=pair.xent_criterion(lis_scores, lis_y)
            # 找到每行的最大值的索引, 将每行最大值所在位置置为1
            max_indices = torch.argmax(lis_scores, dim=1)
            lis_pred = torch.zeros_like(lis_scores)
            # 将每行最大值的索引处的元素置为1
            lis_pred[torch.arange(lis_pred.size(0)), max_indices] = 1
            per_game_acc = 'nan'
            # print('loss',this_loss)
            num_identical_rows = 0
            for row1, row2 in zip(lis_pred, lis_y):
                if torch.equal(row1, row2):
                    num_identical_rows += 1
            this_acc=num_identical_rows/batch_size
            # print('accuracy:',this_acc)

        # Save language
        if args.use_lang:   # sample.py执行
            lang_i = lang1.argmax(2)
            # print(lang_i)
            lang_text_unjoined = util.to_emergent_text(lang_i)
            # print(lang_text_unjoined)
            lang_text = [" ".join(toks) for toks in lang_text_unjoined]
            # print(lang_text)
        else:
            lang_text_unjoined = [["N/A"] for _ in range(batch_size)]
            lang_text = ["N/A" for _ in range(batch_size)]
        true_lang_text = get_true_lang(batch, dataloader.dataset, args, join=False)
        true_lang_text_joined = [" ".join(t) for t in true_lang_text]

        md = torch.tensor([0, 0, 0])
        # Game difficulty/other metadata indicator
        all_lang.extend(zip(lang_text, true_lang_text, per_game_acc, md.numpy()))

        # Get attributes
        all_toks.extend(lang_text_unjoined)

        if args.joint_training:
            # Also train speaker on classification task
            spk_scores = pair.speaker.classify_from_states(states, lis_inp)
            spk_loss = pair.bce_criterion(spk_scores, lis_y)
            spk_pred = (spk_scores > 0).float()
            spk_per_game_acc = (spk_pred == lis_y).float().mean(1).cpu().numpy()
            spk_acc = spk_per_game_acc.mean()
            stats.update(spk_loss=spk_loss, spk_acc=spk_acc)
            comb_loss = this_loss + args.joint_training_lambda * spk_loss
        else:   # sample.py执行
            comb_loss = this_loss

        if training:   # sample.py 不执行
            comb_loss.backward()
            if (batch_i + 1) % args.accum_steps == 0:
                torch.nn.utils.clip_grad_norm_(pair.parameters(), args.clip)
                gcn_opimizer_s.step()
                gcn_opimizer_s.zero_grad()
                gaps_optimizer.step()
                gaps_optimizer.zero_grad()
                gapl_optimizer.step()
                gapl_optimizer.zero_grad()
                gcn_optimizer_l.step()
                gcn_optimizer_l.zero_grad()
                if epoch >= 0:
                    optimizer.step()
                    optimizer.zero_grad()
                backpropped = True
            else:
                backpropped = False
            if batch_i % args.log_interval == 0:
                log_epoch_progress(epoch, batch_i, batch_size, dataloader, stats)

        stats.update(
            loss=this_loss, acc=this_acc, batch_size=batch_size, combined_loss=comb_loss
        )

    if training and not backpropped:
        torch.nn.utils.clip_grad_norm_(pair.parameters(), args.clip)
        optimizer.step()
        optimizer.zero_grad()

    # Compute metrics + collect generation language
    metrics = stats.averages()
    all_lang = pd.DataFrame.from_records(
        all_lang,
        columns=["lang", "true_lang", "acc", "md"],
    )

    if args.use_lang:
        # Compute emergent communication statistics
        # TODO - this should generally be a "meaning preprocess" function
        # if dataloader.dataset.name == "cub":
        #     attrs_numeric = dataloader.dataset.attr_to_numeric(all_attrs)
        # else:
        #     attrs_numeric = None

        lang_metrics = compute_lang_metrics(
            all_lang,
            dataloader.dataset,
            args,
            # attrs=all_attrs,
            # reprs=all_reprs,
            # attrs_numeric=attrs_numeric,
            attrs=None,
            reprs=None,
            attrs_numeric=None,
            toks=all_toks,
        )
        metrics.update(lang_metrics)

    if dataloader.dataset.name == "shapeworld":
        by_md_metrics = compute_metrics_by_md(
            all_lang, md_vocab=dataloader.dataset.metadata_vocab
        )
        metrics.update(by_md_metrics)

    log_epoch_summary(epoch, split, metrics)

    if args.vis:
        vis.report(
            spk_inp.cpu(),
            spk_y.cpu(),
            lis_inp.cpu(),
            lis_y.cpu(),
            dataloader.dataset,
            epoch,
            split,
            {"speaker": lang_text},
            true_lang_text_joined,
            {"speaker": lis_pred},
            exp_dir=os.path.join("exp", args.name),
        )

    clean_language(all_lang)
    # 拼接所有批次的数据
    # 定义填充值，例如 0
    pad_value = 0
    # 对 all_messages 列表中的每个张量进行填充
    for i in range(len(all_messages)):
        msg = all_messages[i]
        seq_len = msg.shape[1]
        if seq_len < args.max_lang_length:
            # 计算需要填充的长度
            pad_length = args.max_lang_length - seq_len
            # 在序列维度（dim=1）后面填充
            padding = (0, 0, 0, pad_length)  # (dim=-1, pad_left, pad_right, dim=-2, pad_top, pad_bottom)
            msg_padded = F.pad(msg, padding, value=pad_value)
            all_messages[i] = msg_padded
    all_messages = torch.cat(all_messages, dim=0)  # 形状 (N, seq_len, vocab_size)
    all_states = torch.cat(all_states, dim=0)      # 形状 (N, embedding_dim)

    torch.save(all_messages, f'KGEC_all_messages_{args.vocab_size}_{args.max_lang_length}_translate.pt')
    torch.save(all_states, f'KGEC_all_states_{args.vocab_size}_{args.max_lang_length}_translate.pt')
    return metrics, all_lang


def clean_language(all_lang_df):
    def clean_lang(lang):
        # Startswith/endswith
        if lang.startswith("<s>"):
            lang = lang[3:]
        if lang.endswith("</s>"):
            lang = lang[:-4]
        return lang

    def clean_true_lang(true_lang):
        return " ".join(true_lang[1:-1])

    all_lang_df["lang"] = all_lang_df["lang"].apply(clean_lang)
    all_lang_df["true_lang"] = all_lang_df["true_lang"].apply(clean_true_lang)


if __name__ == "__main__":
    # for various gcn
    metadata_lis = []
    gcn_model_list = []
    gcn_optimizer = []

    import torch.multiprocessing
    torch.multiprocessing.set_start_method('spawn')
    torch.multiprocessing.set_sharing_strategy('file_system')

    args = io_util.parse_args()
    exp_dir = os.path.join("exp", args.name)

    os.makedirs(exp_dir, exist_ok=True)
    util.save_args(args, exp_dir)
    dataloaders = data.loader.load_dataloaders(args)
    model_config = models.builder.build_models(dataloaders, args)
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    this_game_type = data.util.get_game_type(args)
    checkpoint = torch.load('exp/exp/v7_l7_n10_e5/best_model.pt')
    model_config["pair"].load_state_dict(checkpoint)
    run_args = (model_config["pair"], model_config["optimizer"], dataloaders, args)

    # construct GCN_model for speaker
    out_gcn=512
    model_gcn_s=my_GCN.GNN(input_size=512, hidden_channels=512, out_channels=out_gcn)
    model_gcn_s.cuda()
    checkpoint = torch.load('other_checkpoints/7_7_5_best_gcn_s.pt', map_location='cuda:0')
    model_gcn_s.load_state_dict(checkpoint)
    gcn_opimizer_s = torch.optim.Adam(model_gcn_s.parameters(), lr=0.0001)

    # construct GCN_model for listener
    model_gcn_l=my_GCN.GNN(512, 512, out_gcn)
    model_gcn_l.cuda()
    checkpoint = torch.load('other_checkpoints/7_7_5_best_gcn_l.pt', map_location='cuda:0')
    model_gcn_l.load_state_dict(checkpoint)
    gcn_optimizer_l = torch.optim.Adam(model_gcn_l.parameters(), lr=0.0001)

    # Global pooling for GCN (gap)
    # 定义注意力机制  for speaker and listener
    attention_s = nn.Sequential(
        nn.Linear(out_gcn, 1),  # 注意力权重计算
        nn.Softmax(dim=0)  # 对权重进行归一化
    )
    attention_s.cuda()
    checkpoint = torch.load('other_checkpoints/7_7_5_best_att_s.pt', map_location='cuda:0')
    attention_s.load_state_dict(checkpoint)
    gaps_optimizer = optim.Adam(attention_s.parameters(), lr=0.0001)

    attention_l = nn.Sequential(
        nn.Linear(out_gcn, 1),  # 注意力权重计算
        nn.Softmax(dim=0)  # 对权重进行归一化
    )
    attention_l.cuda()
    checkpoint = torch.load('other_checkpoints/7_7_5_best_att_l.pt', map_location='cuda:0')
    attention_l.load_state_dict(checkpoint)
    gapl_optimizer = optim.Adam(attention_l.parameters(), lr=0.0001)

    all_metrics = []
    metrics = init_metrics()

    for epoch in range(args.epochs):
        # No reset on epoch 0, but reset after epoch 2, epoch 4, etc
        if (
            args.listener_reset_interval > 0
            and (epoch % args.listener_reset_interval) == 0
        ):
            logging.info(f"Resetting listener at epoch {epoch}")
            model_config["pair"].listener.reset_parameters()
        metrics["epoch"] = epoch

        # Train
        # train_metrics, lang = run(0.1, "train", epoch,  *run_args)

        # util.update_with_prefix(metrics, train_metrics, "train")

        # when testting, enabling the code below, Comment out the code above in Train...
        # Test start
        acc_list = []
        nc = 1
        for i in range(51):
            noise = 0.0 +i*0.02
            print('Noise level is:', noise)
            sname = 'test_ref'
            train_metrics, lang = run(noise, sname, epoch, *run_args)
            pdb.set_trace()
            acc = train_metrics['acc']
            print('the accuracy is:', acc)
            acc_list.append(acc)
            pdb.set_trace()
        print(acc_list)

        # 检查 Excel 文件是否存在
        file_name = 'what.xlsx'
        if not os.path.exists(file_name):
            # 文件不存在时，创建一个新的 Excel 文件
            wb = Workbook()
            ws = wb.active
        else:
            # 文件存在时，加载已有的 Excel 文件
            wb = load_workbook(file_name)
            ws = wb.active

        # 找到第一个空列
        nc = ws.max_column + 1

        # 将数据写入第一个空列
        for i, item in enumerate(acc_list, start=9):
            ws.cell(row=i, column=nc, value=item)

        # 保存 Excel 文件
        wb.save(file_name)
        nc = nc+1
        util.update_with_prefix(metrics, train_metrics, "train")
        # Test end

        # Eval across seen/unseen splits, and all game configurations
        for game_type in ["ref"]:
            if args.no_cross_eval and game_type != this_game_type:
                continue
            for split in ["val", "test"]:
                split_metrics = defaultdict(list)

                for split_type in ["", "_same"]:
                    sname = f"{split}{split_type}_{game_type}"
                    if sname in dataloaders:
                        eval_metrics, eval_lang = run(0, sname, epoch, *run_args)    # val_same_ref
                        # eval_metrics, eval_lang = run("train", epoch,  *run_args)
                        util.update_with_prefix(metrics, eval_metrics, sname)
                        if this_game_type == game_type:
                            # Default
                            util.update_with_prefix(
                                metrics, eval_metrics, f"{split}{split_type}"
                            )

                        for metric, value in eval_metrics.items():
                            split_metrics[metric].append(value)

                    if sname == f"test_{this_game_type}":
                        # Store + concatenate test language
                        lang = pd.concat((lang, eval_lang), axis=0)

                # Average across seen and novel
                split_metrics = {k: np.mean(v) for k, v in split_metrics.items()}
                util.update_with_prefix(
                    metrics, split_metrics, f"{split}_avg_{game_type}"
                )
                if this_game_type == game_type:
                    # Default
                    util.update_with_prefix(metrics, split_metrics, f"{split}_avg")

        # Use validation accuracy to choose the best model.
        is_best = metrics["val_avg_acc"] > metrics["best_acc"]
        if is_best:
            metrics["best_acc"] = metrics["val_avg_acc"]
            metrics["best_loss"] = metrics["val_avg_loss"]
            metrics["best_epoch"] = epoch
            if args.use_lang:
                lang.to_csv(os.path.join(exp_dir, "best_lang.csv"), index=False)
            # Save the model
            model_fname = os.path.join(exp_dir, "best_model.pt")
            torch.save(model_config["pair"].state_dict(), model_fname)
            torch.save(model_gcn_l.state_dict(), f'{args.vocab_size}_{args.max_lang_length}_{args.n_examples}_best_gcn_l.pt')
            torch.save(model_gcn_s.state_dict(), f'{args.vocab_size}_{args.max_lang_length}_{args.n_examples}_best_gcn_s.pt')
            torch.save(attention_l.state_dict(), f'{args.vocab_size}_{args.max_lang_length}_{args.n_examples}_best_att_l.pt')
            torch.save(attention_s.state_dict(), f'{args.vocab_size}_{args.max_lang_length}_{args.n_examples}_best_att_s.pt')

        if epoch % args.save_interval == 0:
            model_fname = os.path.join(exp_dir, f"{epoch}_model.pt")
            torch.save(model_config["pair"].state_dict(), model_fname)
            if args.use_lang:
                lang.to_csv(os.path.join(exp_dir, f"{epoch}_lang.csv"), index=False)

        # Additionally track best for splits separately
        metrics["best_val_acc"] = max(metrics["best_val_acc"], metrics["val_acc"])
        if "val_same_acc" in metrics:
            metrics["best_val_same_acc"] = max(
                metrics["best_val_same_acc"], metrics["val_same_acc"]
            )

        all_metrics.append(metrics.copy())

        if args.wandb:
            import wandb
            wandb.log(metrics)

        pd.DataFrame(all_metrics).to_csv(
            os.path.join(exp_dir, "metrics.csv"), index=False
        )
